import pandas as pd
import numpy as np

import os
from openpyxl import load_workbook

def results_to_excel(TRAIN_CONFIG, accuracy, clf_report, iou_results):
    print("Writting results to excel")
    excel_path = "./trained_models/results.xlsx"

    # Results dataframe
    df_results = pd.DataFrame(columns=["test_number", "audit_time", "results_folder","downscaling_factor", "mean_accuracy", 
                                   "mean_precision", "mean_recall", "mean_f1", "mean_iou",
                                   "epithelium_precision", "epithelium_recall","epithelium_f1", "epithelium_iou",
                                   "blood_vessel_precision", "blood_vessel_recall","blood_vessel_f1", "blood_vessel_iou",
                                   "stroma_precision", "stroma_recall",  "stroma_f1", "stroma_iou",
                                   "adipocytes_precision", "adipocytes_recall","adipocytes_f1", "adipocytes_iou",
                                   ])
    # Configuration dataframe
    df_config = pd.DataFrame(columns=["test_number", "audit_time", "results_folder", "network"
                                    , "n_classes", "batch_size", "epochs", "training_images", "validation_split", "test_split", "learning_rate",
                                    "early_stopping_patience", "lr_on_plateau_patience", "lr_mod_factor", "verbose"])

    wb = load_workbook(excel_path)

    ws_results = wb["results"]
    ws_configuration = wb["configurations"]

    idx = ws_results["A"+str(ws_results.max_row)].value
    last_index = 1 if idx == "test_number" else idx + 1

    df_results.loc[0, "test_number"] = last_index
    df_results.loc[0, "audit_time"] = TRAIN_CONFIG["execution_time"]
    df_results.loc[0, "results_folder"] = TRAIN_CONFIG["results_folder"]
    df_results.loc[0, "downscaling_factor"] = TRAIN_CONFIG["downscaling_factor"]

    # RESULTS DATA
    df_results.loc[0, "mean_accuracy"] = np.round(accuracy*100,2)
    df_results.loc[0, "mean_precision"] = np.round(clf_report["macro avg"]["precision"]*100,2)
    df_results.loc[0, "mean_recall"] = np.round(clf_report["macro avg"]["recall"]*100,2)
    df_results.loc[0, "mean_f1"] = np.round(clf_report["macro avg"]["f1-score"]*100,2)

    df_results.loc[0, "epithelium_precision"] = np.round(clf_report["1"]["precision"]*100,2)
    df_results.loc[0, "epithelium_recall"] = np.round(clf_report["1"]["recall"]*100,2)
    df_results.loc[0, "epithelium_f1"] = np.round(clf_report["1"]["f1-score"]*100,2)


    df_results.loc[0, "blood_vessel_precision"] = np.round(clf_report["2"]["precision"]*100,2)
    df_results.loc[0, "blood_vessel_recall"] = np.round(clf_report["2"]["recall"]*100,2)
    df_results.loc[0, "blood_vessel_f1"] = np.round(clf_report["2"]["f1-score"]*100,2)


    df_results.loc[0, "stroma_precision"] = np.round(clf_report["3"]["precision"]*100,2)
    df_results.loc[0, "stroma_recall"] = np.round(clf_report["3"]["recall"]*100,2)
    df_results.loc[0, "stroma_f1"] = np.round(clf_report["3"]["f1-score"]*100,2)


    df_results.loc[0, "adipocytes_precision"] = np.round(clf_report["4"]["precision"]*100,2)
    df_results.loc[0, "adipocytes_recall"] = np.round(clf_report["4"]["recall"]*100,2)
    df_results.loc[0, "adipocytes_f1"] = np.round(clf_report["4"]["f1-score"]*100,2)

    df_results.loc[0, "mean_iou"] = np.round(iou_results["mean_iou"]*100,2)
    df_results.loc[0, "epithelium_iou"] = np.round(iou_results["epithelium_iou"]*100,2)
    df_results.loc[0, "blood_vessel_iou"] = np.round(iou_results["blood_vessel_iou"]*100,2)
    df_results.loc[0, "stroma_iou"] = np.round(iou_results["stroma_iou"]*100,2)
    df_results.loc[0, "adipocytes_iou"] = np.round(iou_results["adipocytes_iou"]*100,2)
    
    # CONFIG DATA
    df_config.loc[0,"test_number"] = last_index
    df_config.loc[0, "audit_time"] = TRAIN_CONFIG["execution_time"]
    df_config.loc[0, "results_folder"] = TRAIN_CONFIG["results_folder"]
    df_config.loc[0, "network"] = TRAIN_CONFIG["network"]
    df_config.loc[0, "n_classes"] = TRAIN_CONFIG["n_classes"]
    df_config.loc[0, "batch_size"] = TRAIN_CONFIG["batch_size"]
    df_config.loc[0, "epochs"] = TRAIN_CONFIG["epochs"]
   
    df_config.loc[0, "training_images"] = len(os.listdir(os.path.join(TRAIN_CONFIG["training_path"],"images")))
    df_config.loc[0, "validation_split"] = TRAIN_CONFIG["validation_split"]
    df_config.loc[0, "test_split"] = TRAIN_CONFIG["test_split"]
    df_config.loc[0, "learning_rate"] = TRAIN_CONFIG["learning_rate"]
    df_config.loc[0, "early_stopping_patience"] = TRAIN_CONFIG["early_stopping_patience"]
    df_config.loc[0, "lr_on_plateau_patience"] = TRAIN_CONFIG["lr_on_plateau_patience"]
    df_config.loc[0, "lr_mod_factor"] = TRAIN_CONFIG["lr_mod_factor"]
    df_config.loc[0,"verbose"] = TRAIN_CONFIG["verbose"]

    ws_results.append(df_results.iloc[0].values.tolist())
    ws_configuration.append(df_config.iloc[0].values.tolist())
    # Save results in excel file
    wb.save(excel_path)
    print(f"Results from the execution saved to {excel_path} successfully")